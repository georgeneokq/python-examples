from openpyxl import load_workbook

def get_food_info():
  """
  Excelファイルを読み込んで、食事の情報を返す
  """
  data_path = 'data/shokuji.xlsx'
  wb = load_workbook(filename=data_path)
  sheet = wb['Sheet1']

  rows = [row for row in sheet.iter_rows(values_only=True)]
  headers = rows[0]
  data = rows[1:]

  # Return result as list of dictionaries
  result = [
    { headers[i]: item[i] for i in range(len(headers)) } for item in data
  ]

  wb.close()
  return result

def test():
  food_info = get_food_info()
  assert food_info is not None

if __name__ == '__main__':
  test()